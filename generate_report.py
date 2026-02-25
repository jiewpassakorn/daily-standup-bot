import json
import os
import re
import sys
from copy import copy
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule

SOURCE_FILE = os.getenv("SOURCE_FILE", "JIRA_Export.xlsx")
OUTPUT_FILE = os.getenv("OUTPUT_FILE", "JIRA_Report.xlsx")
SHEET_NAME = os.getenv("JIRA_SHEET_NAME", "JIRA Data")
TEAM_FILE = Path(__file__).parent / "team.json"

# Load team members from external config
if not TEAM_FILE.is_file():
    print(f"Error: {TEAM_FILE} not found. Copy team.json.example and fill in your team.")
    sys.exit(1)
TEAM_MEMBERS = [
    (m["display"], m["jira_name"])
    for m in json.loads(TEAM_FILE.read_text())
]

JIRA_BASE_URL = os.getenv("JIRA_BASE_URL", "https://yourteam.atlassian.net") + "/browse/"
MAX_DATA_ROWS = 500  # pre-fill formulas for 500 rows

# ── Styles ──────────────────────────────────────────────
header_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
center_align = Alignment(horizontal="center", vertical="center")
thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

STATUS_COLORS = {
    "Closed": "92D050",     # Green
    "Resolved": "00B0F0",   # Blue
    "Open": "FFC000",       # Yellow
    "On Hold": "FF5050",    # Red
}

# ── Column layout for Report sheet ──────────────────────
#  A: (hidden) Row# lookup via FILTER – finds nth valid source row
#  B: Sr. No. (Key)
#  C: Issue Type
#  D: Task (Summary)
#  E: Start Date (Created, date only)
#  F: End Date (Due Date, date only)
#  G..M: Person columns (7 members)
#  N: JIRA URL
ROW_LOOKUP_COL = 1
PERSON_START_COL = 7
PERSON_END_COL = PERSON_START_COL + len(TEAM_MEMBERS) - 1
URL_COL = PERSON_END_COL + 1
MAX_SRC_ROW = MAX_DATA_ROWS + 1  # source rows: 2 .. MAX_SRC_ROW

# Sheet name reference (quoted for Excel formula)
SRC = f"'{SHEET_NAME}'"


def extract_hyperlink_text(formula):
    """Extract display text from =HYPERLINK("url","text") formula."""
    if isinstance(formula, str) and formula.startswith("=HYPERLINK("):
        m = re.search(r',\s*"([^"]+)"\s*\)', formula)
        if m:
            return m.group(1)
    return formula


def main():
    # ── Read source workbook ────────────────────────────────
    source_wb = openpyxl.load_workbook(SOURCE_FILE)
    source_ws = source_wb[SHEET_NAME]

    # ── Create output workbook ──────────────────────────────
    wb = openpyxl.Workbook()

    # ════════════════════════════════════════════════════════
    #  Sheet 1: JIRA Data  (paste raw JIRA export here)
    # ════════════════════════════════════════════════════════
    ws_data = wb.active
    ws_data.title = SHEET_NAME
    ws_data.sheet_properties.tabColor = "4472C4"

    for row in source_ws.iter_rows(min_row=1, values_only=False):
        for cell in row:
            value = cell.value
            # Convert HYPERLINK formulas to plain text for clean references
            value = extract_hyperlink_text(value)
            new_cell = ws_data.cell(row=cell.row, column=cell.column, value=value)
            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.alignment = copy(cell.alignment)
                new_cell.number_format = cell.number_format

    # Auto-fit approximate column widths for JIRA Data
    jira_col_widths = {1: 12, 2: 18, 3: 65, 4: 25, 5: 25, 6: 10, 7: 10, 8: 12, 9: 18, 10: 18, 11: 14}
    for col, w in jira_col_widths.items():
        ws_data.column_dimensions[get_column_letter(col)].width = w

    # ════════════════════════════════════════════════════════
    #  Sheet 2: Report  (all formulas, auto-updates)
    # ════════════════════════════════════════════════════════
    ws_report = wb.create_sheet("Report")
    ws_report.sheet_properties.tabColor = "00B050"

    # ── Row 1: Hidden row with JIRA assignee full names ────
    for i, (_, jira_name) in enumerate(TEAM_MEMBERS):
        ws_report.cell(row=1, column=PERSON_START_COL + i, value=jira_name)
    ws_report.row_dimensions[1].hidden = True

    # ── Row 2: Headers ─────────────────────────────────────
    fixed_headers = ["Sr. No.", "Issue Type", "Task", "Start Date", "End Date"]
    for i, h in enumerate(fixed_headers, start=2):  # start at col B (col A is hidden helper)
        cell = ws_report.cell(row=2, column=i, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    for i, (display, _) in enumerate(TEAM_MEMBERS):
        cell = ws_report.cell(row=2, column=PERSON_START_COL + i, value=display)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    cell = ws_report.cell(row=2, column=URL_COL, value="JIRA URL")
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = thin_border

    # Hide column A (row lookup helper)
    ws_report.column_dimensions["A"].hidden = True

    # ── Row 3+: Formulas ───────────────────────────────────
    # Strategy: FILTER + INDEX (Google Sheets compatible)
    #   - Column A: FILTER finds valid row numbers (non-blank AND non-Dropped)
    #   - INDEX picks the kth valid row
    #   - Other columns use INDEX with $A reference
    #
    # FILTER source ranges
    src_b = f"{SRC}!$B$2:$B${MAX_SRC_ROW}"
    src_h = f"{SRC}!$H$2:$H${MAX_SRC_ROW}"
    src_row_range = f"ROW({SRC}!$B$2:$B${MAX_SRC_ROW})"

    for i in range(MAX_DATA_ROWS):
        r = i + 3  # report row (data starts at row 3)

        # A: Row# lookup (hidden)
        # FILTER returns array of valid row numbers, INDEX picks the kth one
        # ROW()-2 gives k=1 for row 3, k=2 for row 4, etc.
        row_formula = (
            f"=IFERROR(INDEX(FILTER({src_row_range},"
            f"{src_b}<>\"\","
            f"{src_h}<>\"Dropped\"),"
            f"ROW()-2),\"\")"
        )
        ws_report.cell(row=r, column=ROW_LOOKUP_COL, value=row_formula)

        # All other columns use INDEX with $A (the looked-up row number)
        ref = f"$A{r}"

        # B: Sr. No. (= Key)
        ws_report.cell(row=r, column=2, value=f'=IF({ref}="","",INDEX({SRC}!B:B,{ref}))')

        # C: Issue Type
        ws_report.cell(row=r, column=3, value=f'=IF({ref}="","",INDEX({SRC}!A:A,{ref}))')
        ws_report.cell(row=r, column=3).alignment = center_align

        # D: Task (= Summary)
        ws_report.cell(row=r, column=4, value=f'=IF({ref}="","",INDEX({SRC}!C:C,{ref}))')

        # E: Start Date (= Created, date only)
        ws_report.cell(row=r, column=5, value=f'=IF({ref}="","",INT(INDEX({SRC}!I:I,{ref})))')
        ws_report.cell(row=r, column=5).number_format = "D-MMM"
        ws_report.cell(row=r, column=5).alignment = center_align

        # F: End Date (= Due Date, date only)
        ws_report.cell(row=r, column=6, value=f'=IF({ref}="","",INT(INDEX({SRC}!K:K,{ref})))')
        ws_report.cell(row=r, column=6).number_format = "D-MMM"
        ws_report.cell(row=r, column=6).alignment = center_align

        # G–M: Person columns
        # Logic: if Assignee matches this person → show Status
        for j in range(len(TEAM_MEMBERS)):
            col = PERSON_START_COL + j
            col_letter = get_column_letter(col)
            formula = (
                f'=IF({ref}="",'
                f'"",'
                f"IF(INDEX({SRC}!$D:$D,{ref})={col_letter}$1,"
                f"INDEX({SRC}!$G:$G,{ref}),"
                f'""))'
            )
            cell = ws_report.cell(row=r, column=col, value=formula)
            cell.alignment = center_align

        # N: JIRA URL (clean, no tracking params)
        ws_report.cell(
            row=r, column=URL_COL,
            value=f'=IF({ref}="","","{JIRA_BASE_URL}"&INDEX({SRC}!B:B,{ref}))',
        )

        # Apply borders to visible columns (B onwards)
        for col_idx in range(2, URL_COL + 1):
            ws_report.cell(row=r, column=col_idx).border = thin_border

    # ── Column widths ──────────────────────────────────────
    report_widths = {
        2: 18,    # Sr. No.
        3: 12,    # Issue Type
        4: 65,    # Task
        5: 12,    # Start Date
        6: 12,    # End Date
        URL_COL: 50,
    }
    for col, w in report_widths.items():
        ws_report.column_dimensions[get_column_letter(col)].width = w
    for j in range(len(TEAM_MEMBERS)):
        ws_report.column_dimensions[get_column_letter(PERSON_START_COL + j)].width = 16

    # ── Freeze panes (freeze header row) ──────────────────
    ws_report.freeze_panes = "B3"

    # ── Auto filter ────────────────────────────────────────
    ws_report.auto_filter.ref = f"B2:{get_column_letter(URL_COL)}2"

    # ── Conditional formatting for status colors ───────────
    person_range = (
        f"{get_column_letter(PERSON_START_COL)}3:"
        f"{get_column_letter(PERSON_END_COL)}{MAX_DATA_ROWS + 2}"
    )
    for status, color in STATUS_COLORS.items():
        ws_report.conditional_formatting.add(
            person_range,
            CellIsRule(
                operator="equal",
                formula=[f'"{status}"'],
                fill=PatternFill(start_color=color, end_color=color, fill_type="solid"),
            ),
        )

    # ════════════════════════════════════════════════════════
    #  Save
    # ════════════════════════════════════════════════════════
    wb.save(OUTPUT_FILE)
    print(f"Generated: {OUTPUT_FILE}")
    print(f"  - Sheet '{SHEET_NAME}': raw data ({source_ws.max_row - 1} tickets)")
    print(f"  - Sheet 'Report': formula-based ({MAX_DATA_ROWS} rows pre-filled)")
    print(f"  - {len(TEAM_MEMBERS)} team members, {len(STATUS_COLORS)} status colors")


if __name__ == "__main__":
    main()
